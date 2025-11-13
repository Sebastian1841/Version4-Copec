import os
from openpyxl import load_workbook

# Mapeo de nombres conocidos → estandarizados
COLUMN_MAPPING = {
    'fecha': 'Fecha',
    'fecha transacción': 'Fecha Transacción',
    'estación': 'Estación',
    'estación de servicio': 'Estación de Servicio',
    'comuna': 'Comuna',
    'volumen': 'Volumen',
    'litros': 'Litros',
    'precio': 'Precio',
    'monto': 'Monto',
    'patente': 'Patente',
    'tarjeta': 'Tarjeta',
    'odómetro (kms.)': 'Odómetro (Kms.)',
    'odometro': 'Odometro',
    'n° vehículo': 'N° Vehículo',
    'movimiento': 'Movimiento',
    'departamento': 'Departamento',
    'región': 'Región',
    'hora transacción': 'Hora Transacción',
    'guía de despacho': 'Guía Despacho',
    'rut': 'Rut',
    'rut chofer': 'Rut Chofer',
    'rut atendedor': 'Rut Atendedor',
    'rendimiento (kms. por litro)': 'Rendimiento (Kms. por Litro)',
    'rendimiento ($ por km.)': 'Rendimiento ($ por Km.)',
    'versión': 'Versión',
    'nickname': 'Nickname'
}

def process_excel(file_path):
    """
    Lee un .xlsx, detecta la fila de headers y convierte todas las hojas en JSON.
    NO formatea datetime/time aquí. Devuelve valores crudos.
    """
    if not os.path.exists(file_path):
        print(f"❌ Archivo no encontrado: {file_path}")
        return []

    wb = load_workbook(file_path, data_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.values)
        if len(rows) < 2:
            continue

        # Detecta header con 'patente' y 'fecha'/'fecha transacción'
        header_row = None
        for row in rows:
            if not row:
                continue
            lower_cells = [str(c).lower() if c else "" for c in row]
            if any("patente" in c for c in lower_cells) and any(("fecha transacción" in c) or ("fecha" in c) for c in lower_cells):
                header_row = row
                break

        if not header_row:
            continue

        header_index = rows.index(header_row)

        # Normaliza headers
        headers_normalized = [
            COLUMN_MAPPING.get(str(h).lower().strip(), str(h).strip()) if h else None
            for h in header_row
        ]

        # Filas de datos
        for data_row in rows[header_index + 1:]:
            obj = {}
            for idx, cell_value in enumerate(data_row):
                column_name = headers_normalized[idx] if idx < len(headers_normalized) else None
                if column_name is not None and cell_value is not None:
                    obj[column_name] = cell_value  # ⬅ sin formatear
            if obj:
                all_data.append(obj)

    return all_data
